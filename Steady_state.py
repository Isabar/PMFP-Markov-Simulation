# -*- coding: utf-8 -*-
"""
Created on Tue Dec 13 15:02:10 2022

@author: baret
"""

import pandas as pd
import genere_sol 
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook



def calc_matrice_transition(Tri,NbF,Proba, Niveaux):
    Q=np.zeros((NbF*2+1,NbF*2+1))


    for i in range (NbF):
        print('Niveaux')
        print(Niveaux)
        print(Proba)
        print(Proba[i,Niveaux[i]])
        Q[2*i,(2*i+1)]=1-Proba[i,Niveaux[i]]
       
        j=np.where(Tri==(i+1))
        #print(j)
        if j[0] < NbF-1:
           # print(2*Tri[j[0]+1])
            Q[2*i,(2*Tri[j[0]+1])-2]=Proba[i,Niveaux[i]]
            
        else :
            Q[2*i,NbF*2]=Proba[i,Niveaux[i]]
        
        Q[2*(i)+1,(2*i+1)]=1
    Q[NbF*2,NbF*2]=1
    print(Q)
    return Q
    
def calc_steady_state(Q,Tri,NbF):
        
    state0=np.zeros((NbF*2)+1)
    state0[(2*(Tri[0]-1))]=1
   # print('state 0')
   # print(state0)
   # print(Q)
    state1=state0.dot(Q)
    gap=0
    n=1
    staten=np.zeros((NbF*2)+1)
    
    while(gap <5) :
    
        n=n+1
        Pn=np.linalg.matrix_power(Q,n)
        stateN=state0.dot(Pn)
        test=True
        #print(stateN)
        #print(staten)
        for i in range(0,(NbF*2)+1) :
            if stateN[i] <= staten[i]-0.01 :
                
               # print(stateN[i])
                #print(staten[i])
                test =False
            if stateN[i] >= staten[i]+0.01:
               # print(stateN[i])
                #print(staten[i])
                test =False
          #  print(test)
        if test==True:
            gap=gap+1
        staten=stateN
    
    
   # print(staten)
    return staten

def write_steady_state(filename,steady):
    workbook = load_workbook(filename)
    worksheetP=workbook['Proba']
    Size=len(steady)
   # print(Size)
    for k in range (Size):
        Size2=len(steady[k])
        for j in range(Size2):
             c=worksheetP.cell(row=k+1,column=j+1)
             c.value=steady[k][j]
    worksheetP.delete_cols(1)
    worksheetP.delete_cols(2)
    worksheetP.delete_cols(3)
    workbook.save(filename)

def modif_proba_facilities(filename,nbF,nbL):
    workbook = load_workbook(filename)
    worksheetP=workbook['Proba_facilities']
    for k in range(nbF):
        for l in range(nbL):
            if l==0:
                c=worksheetP.cell(row=k+2,column=l+2)
                c.value=0.3
            elif l==1: 
                c=worksheetP.cell(row=k+2,column=l+2)
                c.value=0.2
            elif l==2:
                c=worksheetP.cell(row=k+2,column=l+2)
                c.value=0.1
    workbook.save(filename)
     
def main(directory,NbC,NbF,instance_size,NbL,Niveaux):
    for i in range(instance_size):
        print(f'running instance{i}')
        fichier=directory+f'instance{i}.xlsx'
        
        T=pd.read_excel(fichier,sheet_name="Tri", index_col=0 )
        Tri =T.to_numpy()
      #  print(Tri)
        modif_proba_facilities(fichier, NbF, NbL)
    
        P=pd.read_excel(fichier, sheet_name="Proba_facilities", index_col=0)
        Pr=P.to_numpy()
        Proba=Pr[:,0:NbL]
       # print(Proba)
        Steady_state=[]
        for c in range(NbC):
            Q=calc_matrice_transition(Tri[c], NbF,Proba,Niveaux[i])
            S=calc_steady_state(Q, Tri[c], NbF)
            Steady_state.append(S)
        
        write_steady_state(fichier,Steady_state)
    return Steady_state

NbC=10
NbF=3
nbL=3     
fichier='Sol_possibles.xlsx'   
directory='C:/Users/baret/Documents/GitHub/PMFP-Markov-Simulation/2_niveaux/Instances/'
Niveau=genere_sol.calc_sol()  
genere_sol.print_Sol(fichier, Niveau, 3, 3)
Steady_state=main(directory,NbC,NbF,27,nbL,Niveau )
